import streamlit as st
from streamlit_option_menu import option-menu
from PIL import Image
from streamlit_card import card
import pandas as pd
import os
from datetime import datetime
from alltools import *
import zipfile
from openpyxl import load_workbook

st.set_page_config(page_title="Excel实用工具", page_icon=":full_moon:",layout="wide")
hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            #root > div:nth-child(1) >div >div >div > div >section >div {padding-top:0rem;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)
html_css = """<style> 
                    .font{
                    font-size: 35px;
                    font-family: "Cooper Black";
                    color: #FF9633;
                    </style>
            """
custom_css = """
            <style>
            div.stButton > button {
                font-size: 20px; /* 修改按钮字体大小 */
                padding: 10px 20px; /* 修改按钮内边距 */
                background-color: #4CAF50; /* 修改按钮背景颜色 */
                color: white; /* 修改按钮文字颜色 */
                border: none; /* 去除边框 */
                border-radius: 5px; /* 设置边框圆角 */
            }
            div.stDownloadButton > button {
                font-size: 20px; /* 修改按钮字体大小 */
                padding: 10px 20px; /* 修改按钮内边距 */
                background-color: #4CAF50; /* 修改按钮背景颜色 */
                color: white; /* 修改按钮文字颜色 */
                border: none; /* 去除边框 */
                border-radius: 5px; /* 设置边框圆角 */
            </style>
            """
language_css = """
    <style>
    [data-testid="stFileUploaderDropZone"]  div  div::before {content:"将文件拖放到此处"}
    [data-testid="stFileUploaderDropZone"]  div  div span {display:none}
    [data-testid="stFileUploaderDropZone"]  div  div::after {color:rgba(49,51,63,0.6);font-size:.8em;content:"每个文件限制200MB.XLSX"}
    [data-testid="stFileUploaderDropZone"]  div  div small {display:none}
    [data-testid="stFileUploaderDropZone"][data-testid="baseButton-secondary"] {font-size:0px;}
    [data-testid="stFileUploaderDropZone"][data-testid="baseButton-secondary"]::after {content:"浏览文件";font-size:17px;}
    </style>
    """
with st.sidebar:
    st.image("logo.png",use_column_width=True)
    choose = option-menu(
        "Excel实用工具",
        # ["简介", "sheet自动截图", "分拆sheet","提取PDF表格","提取word表格","提取Excel图片","提取Word图片","Excel表格比对","Excel公式查询"],
        # icons=["house", "images", "microsoft","water", "tree","trophy", "back","stickies", "star","app-indicator"],
        ["简介", "Sheet自动截图", "Sheet自动分拆","提取Excel图片","Excel表格比对","Excel公式查询"],
        icons=["house", "images", "microsoft","trophy", "stickies", "star","app-indicator"],
        menu_icon=":D",default_index=0,
        styles={
            "container": {"padding": "5!important", "background-color": "#fafafa"},
            "icon": {"color": "orange", "font-size": "25px"},
            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px", "--hover-color": "#eee"},
            "nav-link-selected": {"background-color": "green"},}
    )
if choose == "简介":
    col1,col2,col3 = st.columns([1,2,0.5])
    with col2:
        st.markdown(html_css, unsafe_allow_html=True)
        st.markdown('<p class="font" >关于Excel实用小工具集</p>', unsafe_allow_html=True)
       
    # with col1:
    #     logo = Image.open('logo.png')
    #     st.image(logo, width=100)
    # st.markdown('**Ai百宝箱，里面集成了各种工具，欢迎使用**')
    st.success("让Excel办公更Easy，让Life生活更Better",icon=":material/thumb_up:")
    st.error("日常工作中遇到的各种Excel处理难题在这里都可以找到解决方案,五秒钟搞定，让工作生活更轻松，更美满。")

    col1,col2,col3,col4,col5= st.columns([1,1,1,1,1])
    with col1:
        card(
            title='自动截图',
            text='Excel表格截图的自动化处理',
            image='https://pixnio.com/free-images/2024/09/25/2024-09-25-10-54-40-576x864.jpg',
            url= '',
            key='自动截图',
            styles={"card":{'width':'100%','height':'300px'}}
            )
    with col2:    
        card( 
            title='自动分拆', 
            text='Excel表格分拆的自动化处理', 
            image='https://pixnio.com/free-images/2024/09/25/2024-09-25-10-54-40-576x864.jpg',
            url='', 
            key='自动分拆', 
            styles={"card": {'width': '100%', 'height': '300px'}}
            )
    # col1,col2,col3 = st.columns([1,1,1])
    with col3:
        card(
            title='提取图片',
            text='Excel图片提取的自动化处理',
            image='https://pixnio.com/free-images/2024/08/27/2024-08-27-00-28-53-576x1024.jpg',
            url= '',
            key='图片提取',
            styles={"card":{'width':'100%','height':'300px'}}
        )
    with col4:    
        card( 
            title='表格比对', 
            text='Excel表格比对的自动化处理', 
            image='https://pixnio.com/free-images/2024/08/27/2024-08-27-00-28-53-576x1024.jpg',
            url='', 
            key='表格比对', 
            styles={"card": {'width': '100%', 'height': '300px'}} 
        )
    with col5:
        card(
            title='公式查询',         
            text='Excel公式AI查询处理',        
            image='https://pixnio.com/free-images/2024/08/27/2024-08-27-00-28-53-576x1024.jpg',        
            url='',        
            key='AI公式查询',        
            styles={"card": {'width': '100%', 'height': '300px'}}    
        )
elif choose == "Sheet自动截图":
     col1,col2,col3= st.columns([1,2,1])
     with col2:
        st.markdown(html_css, unsafe_allow_html=True)
        st.markdown('<p class="font">sheet自动截图</p>', unsafe_allow_html=True)
     st.success("轻松实现表格数据截图的自动化处理，无论文件中有多少个Sheet表格需要截图，只需点击一下即可！3秒钟生成对应的截图文件。")
     st.error("身为打工人，每天都需要汇总当天公司运营情况，一张一张地将表格截图发微信给老板看，不但效率低下而且容易出错，每当有新的表格需要汇总时，都需要重新进行截图操作，OMG，重复劳动的噩梦。")
     ExcelShotool_dirs = os.path.basename(create_directory('excelshotool'))
     filenameShot = None
     def upload_shot_Save():
        global filenameShot
        style_language_uploader()
        uploaded_files = st.file_uploader("请选择待截图的文件:",accept_multiple_files =True, type=["xlsx"])
        st.markdown(language_css,unsafe_allow_html=True)
        #保存文件
        if uploaded_files:
            for uploaded_file in uploaded_files:
                file_contents = uploaded_file.getvalue()
                file_path = os.path.join(ExcelShotool_dirs, uploaded_file.name)
                filenameShot = uploaded_file.name
                    # 将文件保存到本地文件系统
                with open(file_path, "wb") as f:
                        f.write(file_contents)
                    # 获取文件路径
                xlsx = os.path.join(os.path.dirname(os.path.abspath(__file__)),file_path)
                wb = load_workbook(xlsx)
                sheet_names = wb.sheetnames
                out_img(xlsx,sheet_names,ExcelShotool_dirs)
                st.markdown(custom_css,unsafe_allow_html=True)
                if st.button("Excel表格截图"):
                        with zipfile.ZipFile(f'{filenameShot}shot.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:
                                # 遍历文件夹中的所有文件和子文件夹
                            for root, dirs, files in os.walk(ExcelShotool_dirs):
                                for file in files:
                                    file_path = os.path.join(root, file)
                                    # 将文件添加到压缩文件中
                                    zipf.write(file_path, os.path.relpath(file_path, ExcelShotool_dirs))
                        st.write("提取完成")
                        download_zip_file(f'{filenameShot}shot.zip')
        else:
            return None
     upload_shot_Save()

elif choose == "Sheet自动分拆":
    col1,col2,col3 = st.columns([1,2,0.5])
    with col2:
        st.markdown(html_css, unsafe_allow_html=True)
        st.markdown('<p class="font">Sheet自动分拆</p>', unsafe_allow_html=True)
    st.success("替代手工移动到新 工作簿功能，批量实现将sheet工作表另存为工作簿，早半天下班。")
    st.error("适合处理需要将一个包含多张报表或数据汇总的工作簿快速分解为多个单独文件的场景，提高处理Excel数据的效率，适用于需要对Excel数据进行独立处理或归档场景。")
    tearexceltool_dirs = os.path.basename(create_directory('tearexceltool'))
    filenameTear = None
    def upload_tear_Save():
        global filenameTear
        style_language_uploader()
        uploaded_files = st.file_uploader("请选择待分拆的文件:",accept_multiple_files =True, type=["xlsx"])
        st.markdown(language_css,unsafe_allow_html=True)
        #保存文件
        if uploaded_files:
            for uploaded_file in uploaded_files:
                file_contents = uploaded_file.getvalue()
                file_path = os.path.join(tearexceltool_dirs, uploaded_file.name)
                filenameTear = uploaded_file.name
                    # 将文件保存到本地文件系统
                with open(file_path, "wb") as f:
                        f.write(file_contents)
                    # 获取文件路径
                xlsl = os.path.join(os.path.dirname(os.path.abspath(__file__)),file_path)
                split_data(xlsl,tearexceltool_dirs)
                st.markdown(custom_css,unsafe_allow_html=True)
                if st.button("Excel表格分拆"):
                        with zipfile.ZipFile(f'{filenameTear}.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:
                                # 遍历文件夹中的所有文件和子文件夹
                            for root, dirs, files in os.walk(tearexceltool_dirs):
                                for file in files:
                                    file_path = os.path.join(root, file)
                                    # 将文件添加到压缩文件中
                                    zipf.write(file_path, os.path.relpath(file_path, tearexceltool_dirs))
                        st.write("提取完成")
                        download_zip_file(f'{filenameTear}.zip')
        else:
            return None
    upload_tear_Save()
elif choose == "提取Excel图片":
    col1,col2,col3 = st.columns([1,2,0.5])
    with col2:
        st.markdown(html_css, unsafe_allow_html=True)
        st.markdown('<p class="font">提取Excel图片</p>', unsafe_allow_html=True)
    st.success("日常办公中，经常需要在Excel表格里插入或者导出图片，特别是在制作产品目录、报告或展示数据时。")
    st.error("需要从单个sheet的Excel文件中提取图片，比如为了备份、进一步处理或在其他应用程序中使用。")   
    takeexcelpictool_dirs = os.path.basename(create_directory('excelpictool'))
    filenameTakepic = None
    def upload_takepic_Save():
        global filenameTakepic
    # 选择文件
        style_language_uploader()
        uploaded_files = st.file_uploader("请选择待提取图片的文件:",accept_multiple_files =True, type=["xlsx"])
        st.markdown(language_css,unsafe_allow_html=True)
        # 保存文件
        if uploaded_files:
            for uploaded_file in uploaded_files:
                file_contents = uploaded_file.getvalue()
                file_path = os.path.join(takeexcelpictool_dirs, uploaded_file.name)
                filenameTakepic = uploaded_file.name
                    # 将文件保存到本地文件系统
                with open(file_path, "wb") as f:
                        f.write(file_contents)
                    # 获取文件路径
                # st.write(f"文件地址: {file_path}")
                xlsl = os.path.join(os.path.dirname(os.path.abspath(__file__)),file_path)
                output_dir = takeexcelpictool_dirs
                wb = load_workbook(xlsl)
                ws = wb.active
                with zipfile.ZipFile(xlsl,'r') as zip_ref:
                    #遍历ZIP归档中的所有文件,查找图片文件
                    for zip_info in zip_ref.infolist():
                        if zip_info.filename.endswith(('.png','jpg','.jpeg','.gif','.bmp')):
                            #提取图片到输出目录
                            img_path = os.path.join(output_dir,os.path.basename(zip_info.filename))
                            with zip_ref.open(zip_info) as img_file,open(img_path,'wb') as out_file:
                                out_file.write(img_file.read())
                print(f"Images extracted to {output_dir}")
                st.markdown(custom_css,unsafe_allow_html=True)
                if st.button("表格图片提取"):
                    with zipfile.ZipFile(f'{filenameTakepic}take.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:
                            # 遍历文件夹中的所有文件和子文件夹
                        for root, dirs, files in os.walk(output_dir):
                            for file in files:
                                file_path = os.path.join(root, file)
                                # 将文件添加到压缩文件中
                                zipf.write(file_path, os.path.relpath(file_path, output_dir))
                    st.write("提取完成")
                    download_zip_file(f'{filenameTakepic}take.zip')

        else:
            return None
    upload_takepic_Save()

elif choose == "Excel表格比对":
    col1,col2,col3 = st.columns([1,2,0.5])
    with col2:
        st.markdown(html_css, unsafe_allow_html=True)
        st.markdown('<p class="font">Excel表格比对</p>', unsafe_allow_html=True)
    st.success("在数据处理与分析中，经常需要比较来自不同来源的数据集，特别是处理涉及多个Excel工作簿和工作表的场景时。")
    st.error("工具能有效快速实现两个Excel文件中多个Sheet的数据的比对，找出差异并保存结果，适用于财务审计、数据清洗或者需要跨数据集一致性检查的 场景。")
    import os
    # def download_res(file_path):
    #     if file_path:
    #         #下载
    #         with open(file_path,'rb') as file:
    #             btn = st.download_button(
    #                 label="📥下载文件",
    #                 data=file,
    #                 file_name= file_path.split('/')[-1],
    #                 mime='text/xlsx'
    #             )
    #     else:
    #         print('文件不存在')

    # def delete_empty_folders_in_current_directory():
    #     """
    #     检查当前目录下是否有空的文件夹，并删除这些空文件夹。
    #     """
    #     current_directory = os.getcwd()  # 获取当前目录路径
        
    #     for root, dirs, files in os.walk(current_directory, topdown=False):
    #         for dir_name in dirs:
    #             folder_path = os.path.join(root, dir_name)
    #             if not os.listdir(folder_path):  # 检查文件夹是否为空
    #                 try:
    #                     os.rmdir(folder_path)
    #                     print(f"已删除空文件夹: {folder_path}")
    #                 except Exception as e:
    #                     print(f"删除文件夹 {folder_path} 时出错: {e}")


    # def create_directory(tool): 
    #     delete_empty_folders_in_current_directory()
    #     # 获取当前时间，并格式化为字符串  
    #     current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  
    
    #     # 指定要创建的目录的路径（例如，在当前工作目录下）  
    #     directory_path = os.path.join(os.getcwd(), current_time+tool)  
        
    #     # 创建目录  
    #     try:  
    #         os.mkdir(directory_path)  
    #         print(f"目录 '{directory_path}' 创建成功")  
    #     except OSError as error:  
    #         print(f"创建目录 '{directory_path}' 失败。错误：{error}")
    #     return directory_path
   #directory_name = os.path.basename(create_directory('comparetool'))
    #print(f"目录的名称是: {directory_name}")  
    # 上传文件
    comparetool_dirs = os.path.basename(create_directory('comparetool'))
    filenameCompare = None
    def upload_Save(key):
        global filenameCompare
    # 选择文件
        style_language_uploader()
        uploaded_file = st.file_uploader(f"请选择第{key}个比对文件:",key=key,accept_multiple_files =False, type=["xlsx"])
        # 保存文件
        if uploaded_file:
            file_contents = uploaded_file.getvalue()
            file_path = os.path.join(comparetool_dirs, uploaded_file.name)
            filenameCompare = uploaded_file.name
                # 将文件保存到本地文件系统
            with open(file_path, "wb") as f:
                    f.write(file_contents)
                # 获取文件路径
            # st.write(f"文件地址: {file_path}")
            xlsl = os.path.join(os.path.dirname(os.path.abspath(__file__)),file_path)
            return xlsl 
        else:
            return None
    # col1,col2 = st.columns([1,1])
    # with col1:
    # st.write("请上传第一个文件")
    first_file = upload_Save(1)
    if first_file:
    # 读取 Excel 文件
       xlsx1 = pd.ExcelFile(first_file)
       
    else:
       st.warning("请先上传文件")
       xlsx1 = None
    # with col2:
    # st.write("请上传第二个文件")
    second_file = upload_Save(2)
    if second_file:
    # 读取 Excel 文件
       xlsx2 = pd.ExcelFile(second_file)  
    else:
       st.warning("请先上传文件")
       xlsx2 = None  
    
    if xlsx1 and xlsx2:
    # 读取 Excel 文件中的所有工作表
        sheets1 = {sheet_name: xlsx1.parse(sheet_name) for sheet_name in xlsx1.sheet_names}
        sheets2 = {sheet_name: xlsx2.parse(sheet_name) for sheet_name in xlsx2.sheet_names}

        comparison_results = {}

        for sheet_name in sheets1.keys():
            if sheet_name in sheets2:
                df1 = sheets1[sheet_name]
                df2 = sheets2[sheet_name]
                comparison = df1.merge(df2, how='outer', indicator=True)
                comparison_results[sheet_name] = comparison[comparison['_merge'] != 'both']
        st.markdown(custom_css, unsafe_allow_html=True)
        if st.button("表格比对"):
            with pd.ExcelWriter(f'comparison_results{filenameCompare}.xlsx')  as writer:
                for sheet_name, result in comparison_results.items():
                    if not result.empty:
                        result.to_excel(writer, sheet_name=sheet_name,index = False)
            st.success("比对完成")
            download_res(f'comparison_results{filenameCompare}.xlsx')
        
elif choose == "Excel公式查询":
    col1,col2,col3 = st.columns([1,2,0.5])
    with col2:
        st.markdown(html_css, unsafe_allow_html=True)
        st.markdown('<p class="font">Excel公式查询</p>', unsafe_allow_html=True)
    st.success(" 输入在Excel表格里需要实现的功能场景,然后回车即可得到相应的公式名称以及具体的应用方法,比如:一列里的20个格里的20个姓名放到一个格里去，中间用逗号分隔，在excel中用什么公式能实现？")
    st.error("=TEXTJOIN(', ', TRUE, A1:A20)这里的参数解释如下：', ' 是你想要在姓名之间插入的分隔符。TRUE 表示忽略空单元格。A1:A20 是包含姓名的单元格范围。")
    from openai import OpenAI
 
    client = OpenAI(
        api_key = "sk-IEyLUa4SrhOF2o2J8jKQUVHTViQ7jS4Eh8xsGtoJXLS4oX3x",
        base_url = "https://api.moonshot.cn/v1",
    )
    
    history = [
        {"role": "system", "content": "你是 Kimi，由 Moonshot AI 提供的人工智能助手，你更擅长中文和英文的对话。尤其擅长Excel相关的知识，懂得所有关于Excel表格操作以及对应的公式，并且能够回答任何与Excel有关的问题"}
    ]
    
    def chat(query, history):
        history.append({
            "role": "user", 
            "content": query
        })
        completion = client.chat.completions.create(
            model="moonshot-v1-8k",
            messages=history,
            temperature=0.3,
        )
        result = completion.choices[0].message.content
        history.append({
            "role": "assistant",
            "content": result
        })
        return result
    
    user_query = st.text_input("请输入在Excel里需要实现的功能")
    
    st.markdown(custom_css, unsafe_allow_html=True)

    if st.button("提交"):
        if not user_query:
            st.warning("请输入内容")
        response = chat(user_query,history)
        st.write(f"{response}")
    # if user_query != "":
    #     response = chat(user_query,history)
    #     st.write(f"{response}")
