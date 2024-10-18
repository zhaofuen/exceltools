import os
from datetime import datetime
import streamlit as st
# import xlwings as xw
# import excel3img
from openpyxl import load_workbook, Workbook
from PIL import ImageDraw

def split_data(src, target_folder):
    # # æ£€æŸ¥æºæ–‡ä»¶è·¯å¾„æ˜¯å¦å­˜åœ¨
    # if not os.path.exists(src):
    #     print('æ–‡ä»¶è·¯å¾„ä¸æ­£ç¡®ï¼Œè¯·æ£€æŸ¥')
    #     return
    
    # # ç¡®ä¿ç›®æ ‡æ–‡ä»¶å¤¹å­˜åœ¨
    # os.makedirs(target_folder, exist_ok=True)
    
    try:
        # åŠ è½½å·¥ä½œç°¿
        workbook = load_workbook(filename=src)
        
        for sheet in workbook:
            # å¤„ç†å·¥ä½œè¡¨åç§°ï¼Œé¿å…éæ³•å­—ç¬¦
            safe_name = ''.join([c for c in sheet.title if c.isalpha() or c.isdigit() or c == ' ']).rstrip()
            
            # åˆ›å»ºä¸€ä¸ªæ–°çš„å·¥ä½œç°¿
            new_workbook = Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = sheet.title
            
            # å¤åˆ¶å·¥ä½œè¡¨å†…å®¹
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)
            
            # ä¿å­˜æ–°çš„å·¥ä½œç°¿
            new_workbook_path = os.path.join(target_folder, f"{safe_name}.xlsx")
            new_workbook.save(new_workbook_path)
            print(f"å·¥ä½œè¡¨ {sheet.title} å·²ä¿å­˜åˆ° {new_workbook_path}")
    
    except Exception as e:
        print(f"é”™è¯¯ä¿¡æ¯: {e}")

# ç¤ºä¾‹è°ƒç”¨
# split_data("example.xlsx", "./æŠ¥è¡¨/")
def delete_empty_folders_in_current_directory():
        """
        æ£€æŸ¥å½“å‰ç›®å½•ä¸‹æ˜¯å¦æœ‰ç©ºçš„æ–‡ä»¶å¤¹ï¼Œå¹¶åˆ é™¤è¿™äº›ç©ºæ–‡ä»¶å¤¹ã€‚
        """
        current_directory = os.getcwd()  # è·å–å½“å‰ç›®å½•è·¯å¾„
        
        for root, dirs, files in os.walk(current_directory, topdown=False):
            for dir_name in dirs:
                folder_path = os.path.join(root, dir_name)
                if not os.listdir(folder_path):  # æ£€æŸ¥æ–‡ä»¶å¤¹æ˜¯å¦ä¸ºç©º
                    try:
                        os.rmdir(folder_path)
                        print(f"å·²åˆ é™¤ç©ºæ–‡ä»¶å¤¹: {folder_path}")
                    except Exception as e:
                        print(f"åˆ é™¤æ–‡ä»¶å¤¹ {folder_path} æ—¶å‡ºé”™: {e}")
def create_directory(tool): 
        delete_empty_folders_in_current_directory()
        # è·å–å½“å‰æ—¶é—´ï¼Œå¹¶æ ¼å¼åŒ–ä¸ºå­—ç¬¦ä¸²  
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  
    
        # æŒ‡å®šè¦åˆ›å»ºçš„ç›®å½•çš„è·¯å¾„ï¼ˆä¾‹å¦‚ï¼Œåœ¨å½“å‰å·¥ä½œç›®å½•ä¸‹ï¼‰  
        directory_path = os.path.join(os.getcwd(), current_time+tool)  
        
        # åˆ›å»ºç›®å½•  
        try:  
            os.mkdir(directory_path)  
            print(f"ç›®å½• '{directory_path}' åˆ›å»ºæˆåŠŸ")  
        except OSError as error:  
            print(f"åˆ›å»ºç›®å½• '{directory_path}' å¤±è´¥ã€‚é”™è¯¯ï¼š{error}")
        return directory_path

def download_res(file_path):
        if file_path:
            #ä¸‹è½½
            with open(file_path,'rb') as file:
                btn = st.download_button(
                    label="ğŸ“¥ä¸‹è½½æ–‡ä»¶",
                    data=file,
                    file_name= file_path.split('/')[-1],
                    mime='text/xlsx'
                )
        else:
            print('æ–‡ä»¶ä¸å­˜åœ¨')
def download_zip_file(zip_file_path):
    """
    æä¾›ä¸‹è½½ ZIP æ–‡ä»¶çš„åŠŸèƒ½ã€‚

    :param zip_file_path: ZIP æ–‡ä»¶çš„è·¯å¾„
    """
    # è¯»å– ZIP æ–‡ä»¶å†…å®¹
    with open(zip_file_path, "rb") as f:
        zip_data = f.read()

    # åˆ›å»ºä¸‹è½½æŒ‰é’®
    st.download_button(
        label="ä¸‹è½½ ZIP æ–‡ä»¶",
        data=zip_data,
        file_name=os.path.basename(zip_file_path),
        mime="application/zip"
    )


languages = {
    "CN":{
        "button":"æµè§ˆæ–‡ä»¶",
        "instructions":"å°†æ–‡ä»¶æ‹–æ”¾åˆ°æ­¤å¤„",
        "limits":"æ¯ä¸ªæ–‡ä»¶é™åˆ¶200MB",
    },
    # "EN":{
    #     "button":"Browse Files",
    #     "instructions":"Drag and drop files here",
    #     "limits":"Each file limited to 200MB",
    # },
}

lang = None

def style_language_uploader():
    language = 'CN'
    hide_label = (
        """
    <style>
        div[data-testid="stFileUploader"]>section[data-testid="stFileUploaderDropzone"]>button[data-testid="baseButton-secondary"] {
           color: white;
        }
        div[data-testid="stFileUploader"]>section[data-testid="stFileUploaderDropzone"]>button[data-testid="baseButton-secondary"]::after {
            content: "BUTTON_TEXT";
            color:black;
            display: block;
            position: absolute;
        }
        div[data-testid="stFileUploaderDropzoneInstructions"]>div>span {
           visibility:hidden;
        }
        div[data-testid="stFileUploaderDropzoneInstructions"]>div>span::after {
           content:"INSTRUCTIONS_TEXT";
           visibility:visible;
           display:block;
        }
         div[data-testid="stFileUploaderDropzoneInstructions"]>div>small {
           visibility:hidden;
        }
        div[data-testid="stFileUploaderDropzoneInstructions"]>div>small::before {
           content:"FILE_LIMITS";
           visibility:visible;
           display:block;
        }
    </style>
     """.replace("BUTTON_TEXT", languages.get(language).get("button"))
        .replace("INSTRUCTIONS_TEXT", languages.get(language).get("instructions"))
        .replace("FILE_LIMITS", languages.get(language).get("limits"))
    )

    st.markdown(hide_label, unsafe_allow_html=True)


# def split_data(src,target_folder):    
#     # if not os.path.exists(src):        
#     #     print('æ–‡ä»¶è·¯å¾„ä¸æ­£ç¡®ï¼Œè¯·æ£€æŸ¥')        
#     #     return    
#     # target_folder = './æŠ¥è¡¨/'    
#     # os.makedirs(target_folder, exist_ok=True)    
#     # å¯åŠ¨Excelåº”ç”¨ï¼Œä¸æ˜¾ç¤ºç•Œé¢    
#     app = xw.App(visible=False, add_book=False)    
#     try:        
#             # åŠ è½½å·¥ä½œç°¿        
#             workbook = app.books.open(src)        
#             for sheet in workbook.sheets:            # å¤„ç†å·¥ä½œè¡¨åç§°ï¼Œé¿å…éæ³•å­—ç¬¦            
#                  safe_name = ''.join([c for c in sheet.name if c.isalpha() or c.isdigit() or c == ' ']).rstrip()
#                  workbook_split = app.books.add()            
#                  sheet_split = workbook_split.sheets[0]            
#                  sheet.api.Copy(Before=sheet_split.api)            
#                  workbook_split.save(os.path.join(                
#                  target_folder, f"{safe_name}.xlsx"))            
#                  workbook_split.close()    
#     except Exception as e:        
#         print(f"é”™è¯¯ä¿¡æ¯: {e}")    
#     finally:        # å…³é—­å·¥ä½œç°¿        
#         workbook.close()        # å…³é—­Excelå®ä¾‹        
#         app.quit()

# def batch_process_excel_files(folder_path):    
#      print("æ‰¹é‡å¤„ç†æ–‡ä»¶å¤¹ä¸‹æ‰€æœ‰Excelæ–‡ä»¶...")    # éå†æ–‡ä»¶å¤¹ä¸­çš„æ‰€æœ‰æ–‡ä»¶    
#      for file in os.listdir(folder_path):        
#         if file.endswith(('.xls', '.xlsx', '.xlsm')):           
#              file_path = os.path.join(folder_path, file)            
#              split_data(file_path)

# file_path = './data/'
# batch_process_excel_files(file_path)
# Excel è½¬æ¢ä¸ºå›¾ç‰‡çš„å‡½æ•°
# def out_img(excel_file, sheet_list,outputdir):
#     """
#     å°†Excelæ–‡ä»¶ä¸­çš„æŒ‡å®šå·¥ä½œè¡¨è½¬æ¢ä¸ºå›¾ç‰‡ã€‚

#     å‚æ•°:
#     excel_file: string, Excelæ–‡ä»¶çš„è·¯å¾„ã€‚
#     sheet_list: list, éœ€è¦è½¬æ¢ä¸ºå›¾ç‰‡çš„å·¥ä½œè¡¨åç§°åˆ—è¡¨ã€‚
#     outputdir: string, è¾“å‡ºå›¾ç‰‡æ–‡ä»¶çš„ç›®å½•è·¯å¾„ã€‚
#     æˆªå›¾æ•ˆæœéå¸¸å¥½ã€‚
#     è¿”å›:
#     æ— è¿”å›å€¼ï¼Œä½†ä¼šåœ¨å½“å‰ç›®å½•ä¸‹ç”Ÿæˆå¯¹åº”å·¥ä½œè¡¨çš„å›¾ç‰‡æ–‡ä»¶ã€‚
#     """
#     try:
#         # å¼€å§‹è½¬æ¢æ“ä½œå‰çš„æç¤º
#         print("å¼€å§‹æˆªå›¾ï¼Œè¯·è€å¿ƒç­‰å¾…....")
#         # éå†å·¥ä½œè¡¨åˆ—è¡¨ï¼Œå¯¹æ¯ä¸ªå·¥ä½œè¡¨è¿›è¡Œè½¬æ¢
#         for sheet_name in sheet_list:
#             # è°ƒç”¨excel2imgæ¨¡å—çš„export_imgå‡½æ•°è¿›è¡Œè½¬æ¢
#             # å‚æ•°åŒ…æ‹¬Excelæ–‡ä»¶è·¯å¾„ã€å›¾ç‰‡æ–‡ä»¶åã€å·¥ä½œè¡¨åå’Œè‡ªå®šä¹‰é…ç½®ï¼ˆè¿™é‡Œè®¾ä¸ºNoneï¼‰

#             image_filname = os.path.join(outputdir,f"{sheet_name}.png")
#             excel3img.export_img(excel_file, image_filname,sheet_name, None)
#             print(f"{sheet_name} æˆªå›¾å®Œæˆ")
#     except Exception as e:
#         # æ•è·è½¬æ¢è¿‡ç¨‹ä¸­å¯èƒ½å‡ºç°çš„å¼‚å¸¸ï¼Œå¹¶æ‰“å°å¼‚å¸¸ä¿¡æ¯
#         print("æˆªå›¾å¤±è´¥", e)
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO

def out_img(excel_file, sheet_list, outputdir):
    """
    å°†Excelæ–‡ä»¶ä¸­çš„æŒ‡å®šå·¥ä½œè¡¨è½¬æ¢ä¸ºå›¾ç‰‡ã€‚

    å‚æ•°:
    excel_file: string, Excelæ–‡ä»¶çš„è·¯å¾„ã€‚
    sheet_list: list, éœ€è¦è½¬æ¢ä¸ºå›¾ç‰‡çš„å·¥ä½œè¡¨åç§°åˆ—è¡¨ã€‚
    outputdir: string, è¾“å‡ºå›¾ç‰‡æ–‡ä»¶çš„ç›®å½•è·¯å¾„ã€‚
    è¿”å›:
    æ— è¿”å›å€¼ï¼Œä½†ä¼šåœ¨æŒ‡å®šç›®å½•ä¸‹ç”Ÿæˆå¯¹åº”å·¥ä½œè¡¨çš„å›¾ç‰‡æ–‡ä»¶ã€‚
    æ­¤å‡½æ•°åªèƒ½æå®šäº†å†…å®¹è½¬å­˜å¹¶æˆªå›¾ï¼Œä¸§å¤±äº†åŸæ¥çš„å„ç§æ ¼å¼ã€‚
    """
    try:
        # å¼€å§‹è½¬æ¢æ“ä½œå‰çš„æç¤º
        print("å¼€å§‹æˆªå›¾ï¼Œè¯·è€å¿ƒç­‰å¾…....")
        
        # ç¡®ä¿è¾“å‡ºç›®å½•å­˜åœ¨
        if not os.path.exists(outputdir):
            os.makedirs(outputdir)
        
        # åŠ è½½å·¥ä½œç°¿
        workbook = load_workbook(filename=excel_file)
        
        # éå†å·¥ä½œè¡¨åˆ—è¡¨ï¼Œå¯¹æ¯ä¸ªå·¥ä½œè¡¨è¿›è¡Œè½¬æ¢
        for sheet_name in sheet_list:
            # è·å–æŒ‡å®šçš„å·¥ä½œè¡¨
            sheet = workbook[sheet_name]
            
            # åˆ›å»ºä¸€ä¸ªç©ºç™½çš„å›¾ç‰‡
            img = PILImage.new('RGB', (1000, 1000), color = (255, 255, 255))
            draw = ImageDraw.Draw(img)
            
            # è·å–å·¥ä½œè¡¨çš„è¡Œå’Œåˆ—
            for row in sheet.iter_rows():
                for cell in row:
                    # è·å–å•å…ƒæ ¼çš„å€¼
                    cell_value = cell.value
                    if cell_value is not None:
                        # åœ¨å›¾ç‰‡ä¸Šç»˜åˆ¶å•å…ƒæ ¼å†…å®¹
                        draw.text((cell.column * 100, cell.row * 50), str(cell_value), fill="black")
            
            # ä¿å­˜å›¾ç‰‡
            image_filename = os.path.join(outputdir, f"{sheet_name}.png")
            img.save(image_filename)
            
            print(f"{sheet_name} æˆªå›¾å®Œæˆ")
    
    except Exception as e:
        # æ•è·è½¬æ¢è¿‡ç¨‹ä¸­å¯èƒ½å‡ºç°çš„å¼‚å¸¸ï¼Œå¹¶æ‰“å°å¼‚å¸¸ä¿¡æ¯
        print("æˆªå›¾å¤±è´¥", e)

# ç¤ºä¾‹è°ƒç”¨
# out_img("example.xlsx", ["Sheet1", "Sheet2"], "./output_images")